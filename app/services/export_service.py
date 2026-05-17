"""
Export Service v1.2
===================
Methods:
  export_excel()          → .xlsx with 7 sheets (Invoices, Vendors, Compliance Issues,
                            Anomalies, Manual Review Queue, Line Items, Summary)
  export_sap_csv()        → SAP/ERP-ready semicolon-delimited CSV
  export_powerbi_excel()  → 4-page Power BI-ready .xlsx (Executive, Compliance,
                            Vendor Risk, Manual Review Queue)
  export_powerbi_json()   → denormalised JSON (legacy / kept for compatibility)
"""
from __future__ import annotations
import csv
import json
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Optional

import pandas as pd
from loguru import logger
from sqlalchemy.orm import Session

from app.config import settings
from app.models.invoice import Invoice, LineItem, ComplianceResult, AnomalyFlag


# ── Colour palette ────────────────────────────────────────────────────────────
_BLUE_DARK   = "1F4E79"   # header background
_BLUE_MID    = "2E75B6"   # sub-header
_WHITE       = "FFFFFF"
_GREEN       = "C6EFCE"
_YELLOW      = "FFEB9C"
_RED         = "FFC7CE"
_ORANGE      = "FCE4D6"
_GREY        = "F2F2F2"


class ExportService:
    def __init__(self, db: Session):
        self.db = db
        settings.export_dir.mkdir(parents=True, exist_ok=True)

    # ── Query helpers ─────────────────────────────────────────────────────────

    def _invoices(self, ids=None) -> list[Invoice]:
        q = self.db.query(Invoice)
        if ids:
            q = q.filter(Invoice.id.in_(ids))
        return q.order_by(Invoice.id).all()

    # ── DataFrame builders ────────────────────────────────────────────────────

    def _inv_df(self, invoices: list[Invoice]) -> pd.DataFrame:
        return pd.DataFrame([{
            "Invoice ID":        inv.id,
            "Filename":          inv.original_filename,
            "Invoice Number":    inv.invoice_number,
            "Vendor Name":       inv.vendor_name,
            "Currency":          inv.currency,
            "Total Amount":      inv.total_amount,
            "Tax Amount":        inv.tax_amount,
            "Tax Rate %":        inv.tax_rate_percent,
            "Invoice Date":      inv.invoice_date,
            "Due Date":          inv.due_date,
            "Swiss UID":         inv.swiss_uid,
            "IBAN":              inv.iban,
            "QR Reference":      inv.qr_reference,
            "Language":          inv.language.value if inv.language else None,
            "Status":            inv.status.value if inv.status else None,
            "Compliance":        str(inv.overall_compliance_status.value
                                    if hasattr(inv.overall_compliance_status, "value")
                                    else inv.overall_compliance_status),
            "Anomaly Score":     inv.anomaly_score or 0,
            "Confidence":        inv.extraction_confidence,
        } for inv in invoices])

    def _vendor_df(self, invoices: list[Invoice]) -> pd.DataFrame:
        """Aggregate vendor statistics from processed invoices."""
        vendor_data: dict[str, dict] = {}
        for inv in invoices:
            name = inv.vendor_name or "Unknown"
            if name not in vendor_data:
                vendor_data[name] = {
                    "Vendor Name":      name,
                    "Swiss UID":        inv.swiss_uid,
                    "IBAN":             inv.iban,
                    "Invoice Count":    0,
                    "Total Value":      0.0,
                    "Currencies Used":  set(),
                    "Avg Anomaly Score":0.0,
                    "Scores":           [],
                    "Compliance Fails": 0,
                    "First Seen":       inv.invoice_date,
                    "Last Seen":        inv.invoice_date,
                }
            v = vendor_data[name]
            v["Invoice Count"] += 1
            v["Total Value"] += inv.total_amount or 0
            if inv.currency:
                v["Currencies Used"].add(inv.currency)
            v["Scores"].append(inv.anomaly_score or 0)
            cs = str(inv.overall_compliance_status.value
                     if hasattr(inv.overall_compliance_status, "value")
                     else inv.overall_compliance_status)
            if cs == "fail":
                v["Compliance Fails"] += 1
            # track date range
            if inv.invoice_date:
                if not v["First Seen"] or inv.invoice_date < v["First Seen"]:
                    v["First Seen"] = inv.invoice_date
                if not v["Last Seen"] or inv.invoice_date > v["Last Seen"]:
                    v["Last Seen"] = inv.invoice_date

        rows = []
        for v in vendor_data.values():
            avg_score = round(sum(v["Scores"]) / len(v["Scores"]), 1) if v["Scores"] else 0
            risk = "🔴 High" if avg_score >= 60 else "🟡 Medium" if avg_score >= 30 else "🟢 Low"
            rows.append({
                "Vendor Name":        v["Vendor Name"],
                "Swiss UID":          v["Swiss UID"],
                "IBAN":               v["IBAN"],
                "Invoice Count":      v["Invoice Count"],
                "Total Value":        round(v["Total Value"], 2),
                "Currencies Used":    ", ".join(sorted(v["Currencies Used"])),
                "Avg Anomaly Score":  avg_score,
                "Risk Level":         risk,
                "Compliance Fails":   v["Compliance Fails"],
                "First Invoice Date": v["First Seen"],
                "Last Invoice Date":  v["Last Seen"],
            })
        rows.sort(key=lambda r: r["Avg Anomaly Score"], reverse=True)
        return pd.DataFrame(rows)

    def _compliance_issues_df(self, ids: list[int]) -> pd.DataFrame:
        """Only FAIL and WARNING compliance results."""
        rows = (
            self.db.query(ComplianceResult)
            .filter(
                ComplianceResult.invoice_id.in_(ids),
                ComplianceResult.status.in_(["fail", "warning"]),
            )
            .order_by(ComplianceResult.invoice_id)
            .all()
        )
        return pd.DataFrame([{
            "Invoice ID":     r.invoice_id,
            "Rule ID":        r.rule_id,
            "Rule Name":      r.rule_name,
            "Category":       r.category,
            "Status":         r.status.value if r.status else None,
            "Message":        r.message,
            "Field Checked":  r.field_checked,
            "Actual Value":   r.actual_value,
            "Checked At":     r.checked_at.isoformat() if r.checked_at else None,
        } for r in rows])

    def _anomaly_df(self, ids: list[int]) -> pd.DataFrame:
        rows = (
            self.db.query(AnomalyFlag)
            .filter(AnomalyFlag.invoice_id.in_(ids))
            .order_by(AnomalyFlag.invoice_id)
            .all()
        )
        return pd.DataFrame([{
            "Invoice ID":          r.invoice_id,
            "Anomaly Type":        r.anomaly_type,
            "Severity":            r.severity,
            "Score Contribution":  r.score_contribution,
            "Description":         r.description,
            "Recommended Action":  r.recommended_action,
            "Detected At":         r.detected_at.isoformat() if r.detected_at else None,
        } for r in rows])

    def _manual_review_df(self, invoices: list[Invoice]) -> pd.DataFrame:
        """Invoices needing manual review: anomaly_score ≥ 40 OR compliance = fail."""
        rows = []
        for inv in invoices:
            cs = str(inv.overall_compliance_status.value
                     if hasattr(inv.overall_compliance_status, "value")
                     else inv.overall_compliance_status)
            score = inv.anomaly_score or 0
            if score < 40 and cs != "fail":
                continue
            # Gather top issues
            issues = []
            for flag in (inv.anomaly_flags or []):
                issues.append(f"{flag.anomaly_type} ({flag.severity})")
            for cr in (inv.compliance_results or []):
                if cr.status and cr.status.value in ("fail", "warning"):
                    issues.append(f"{cr.rule_id}: {cr.message or cr.rule_name}")
            severity = "🔴 Critical" if score >= 70 else "🟠 High" if score >= 40 else "🟡 Medium"
            rows.append({
                "Invoice ID":          inv.id,
                "Invoice Number":      inv.invoice_number,
                "Vendor":              inv.vendor_name,
                "Invoice Date":        inv.invoice_date,
                "Total Amount":        inv.total_amount,
                "Currency":            inv.currency,
                "Compliance Status":   cs,
                "Anomaly Score":       score,
                "Severity":            severity,
                "Issues":              " | ".join(issues[:5]) if issues else "—",
                "Recommended Action":  "Manual review required",
            })
        rows.sort(key=lambda r: r["Anomaly Score"], reverse=True)
        return pd.DataFrame(rows)

    def _li_df(self, ids: list[int]) -> pd.DataFrame:
        rows = self.db.query(LineItem).filter(LineItem.invoice_id.in_(ids)).all()
        return pd.DataFrame([{
            "Invoice ID":  r.invoice_id,
            "Position":    r.position,
            "Description": r.description,
            "Quantity":    r.quantity,
            "Unit":        r.unit,
            "Unit Price":  r.unit_price,
            "Total Price": r.total_price,
        } for r in rows])

    # ── openpyxl sheet writer ─────────────────────────────────────────────────

    @staticmethod
    def _write_sheet(ws, df: pd.DataFrame, header_color: str = _BLUE_DARK,
                     status_col: str | None = None):
        """Write a DataFrame to a worksheet with styled headers and auto-width."""
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Alignment, Font, PatternFill

        hfill  = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        hfont  = Font(bold=True, color=_WHITE)
        center = Alignment(horizontal="center")

        # Header row
        for ci, col in enumerate(df.columns, 1):
            c = ws.cell(1, ci, col)
            c.font  = hfill and hfont
            c.fill  = hfill
            c.alignment = center

        # Data rows
        status_ci = None
        if status_col and status_col in df.columns:
            status_ci = list(df.columns).index(status_col) + 1

        for ri, row in enumerate(df.itertuples(index=False), 2):
            for ci, val in enumerate(row, 1):
                ws.cell(ri, ci, val)
            if status_ci:
                cell = ws.cell(ri, status_ci)
                v = str(cell.value or "").lower()
                color = (_RED if v == "fail" else
                         _YELLOW if v in ("warning", "medium") else
                         _GREEN if v in ("pass", "low") else
                         _ORANGE if v in ("high", "critical") else _GREY)
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        # Column auto-width
        for ci, cells in enumerate(ws.columns, 1):
            width = max((len(str(c.value or "")) for c in cells), default=10)
            ws.column_dimensions[get_column_letter(ci)].width = min(width + 2, 55)

    @staticmethod
    def _kv_sheet(ws, title: str, rows: list[tuple[str, object]],
                  header_color: str = _BLUE_DARK):
        """Write a simple key-value table (label | value)."""
        from openpyxl.styles import Font, PatternFill, Alignment
        hfill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        ws.cell(1, 1, title).font = Font(bold=True, size=14, color=_WHITE)
        ws.cell(1, 1).fill = hfill
        ws.merge_cells("A1:B1")
        for ri, (k, v) in enumerate(rows, 2):
            ws.cell(ri, 1, k).font = Font(bold=True)
            ws.cell(ri, 2, v)
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 25

    # ─────────────────────────────────────────────────────────────────────────
    # PUBLIC: Excel full export  (7 sheets)
    # ─────────────────────────────────────────────────────────────────────────

    def export_excel(self, invoice_ids=None, filename=None) -> Path:
        """
        Multi-sheet Excel workbook:
          1. Summary          — KPI overview
          2. Invoices         — all invoice fields
          3. Vendors          — aggregated vendor statistics
          4. Compliance Issues— FAIL/WARNING rows only
          5. Anomalies        — detected anomaly flags
          6. Manual Review    — invoices requiring human review
          7. Line Items       — individual line items
        """
        from openpyxl import Workbook

        invs = self._invoices(invoice_ids)
        if not invs:
            raise ValueError("No invoices found.")
        ids = [i.id for i in invs]

        inv_df    = self._inv_df(invs)
        vendor_df = self._vendor_df(invs)
        comp_df   = self._compliance_issues_df(ids)
        anom_df   = self._anomaly_df(ids)
        review_df = self._manual_review_df(invs)
        li_df     = self._li_df(ids)

        filename = filename or f"invoices_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        path = settings.export_dir / filename
        wb = Workbook()

        # ── Sheet 1: Summary KPIs ──────────────────────────────────────────
        ws_sum = wb.active
        ws_sum.title = "Summary"
        total_val   = sum(i.total_amount or 0 for i in invs)
        pass_count  = sum(1 for i in invs if str(getattr(i.overall_compliance_status, "value",
                          i.overall_compliance_status)) == "pass")
        fail_count  = sum(1 for i in invs if str(getattr(i.overall_compliance_status, "value",
                          i.overall_compliance_status)) == "fail")
        warn_count  = sum(1 for i in invs if str(getattr(i.overall_compliance_status, "value",
                          i.overall_compliance_status)) == "warning")
        high_risk   = sum(1 for i in invs if (i.anomaly_score or 0) >= 60)

        self._kv_sheet(ws_sum, "Swiss Invoice Compliance AI — Export Summary", [
            ("Generated At",           datetime.now().strftime("%Y-%m-%d %H:%M")),
            ("Total Invoices",         len(invs)),
            ("Total Invoice Value",    f"{total_val:,.2f}"),
            ("Compliance — Pass",      pass_count),
            ("Compliance — Warning",   warn_count),
            ("Compliance — Fail",      fail_count),
            ("Pass Rate",              f"{pass_count/len(invs)*100:.1f}%" if invs else "—"),
            ("Anomaly Flags",          len(anom_df)),
            ("High-Risk Invoices (≥60)", high_risk),
            ("Unique Vendors",         vendor_df.shape[0]),
            ("Invoices Needing Review", review_df.shape[0]),
        ])

        # ── Sheet 2-7 ─────────────────────────────────────────────────────
        sheet_specs = [
            ("Invoices",          inv_df,    "Compliance",      _BLUE_DARK),
            ("Vendors",           vendor_df, "Risk Level",      _BLUE_MID),
            ("Compliance Issues", comp_df,   "Status",          _BLUE_DARK),
            ("Anomalies",         anom_df,   "Severity",        _BLUE_MID),
            ("Manual Review",     review_df, "Severity",        _BLUE_DARK),
            ("Line Items",        li_df,     None,              _BLUE_MID),
        ]
        for title, df, scol, hcol in sheet_specs:
            ws = wb.create_sheet(title)
            if df.empty:
                ws["A1"] = f"No {title.lower()} data."
            else:
                self._write_sheet(ws, df, header_color=hcol, status_col=scol)

        wb.save(path)
        logger.info(f"Excel export saved: {path}")
        return path

    # ─────────────────────────────────────────────────────────────────────────
    # PUBLIC: SAP / ERP CSV
    # ─────────────────────────────────────────────────────────────────────────

    def export_sap_csv(self, invoice_ids=None, filename=None) -> Path:
        """
        SAP FI/MM + ERP-ready CSV (semicolon-delimited, UTF-8-BOM).
        Columns useful for direct ERP import:
          vendor_name, invoice_number, invoice_date, due_date, iban, currency,
          total_amount, tax_amount, payment_reference, compliance_status, anomaly_score
          + SAP-specific: COMPANY_CODE, DOC_TYPE, PSTNG_DATE, PMNTTRMS, …
        """
        invs = self._invoices(invoice_ids)
        if not invs:
            raise ValueError("No invoices found.")

        filename = filename or f"sap_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        path = settings.export_dir / filename

        fields = [
            # ERP business fields
            "vendor_name", "invoice_number", "invoice_date", "due_date",
            "iban", "currency", "total_amount", "tax_amount",
            "payment_reference", "compliance_status", "anomaly_score",
            # SAP-specific fields
            "COMPANY_CODE", "DOC_TYPE", "PSTNG_DATE", "PMNTTRMS",
            "REF_DOC_NO", "GROSS_AMOUNT", "CALC_TAX_IND", "TAX_AMOUNT",
            "ITEM_TEXT", "VENDOR_NAME_SAP", "IBAN_SAP", "VAT_REG_NO",
            "SWISS_UID", "QR_REFERENCE",
        ]

        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.DictWriter(f, fieldnames=fields, delimiter=";")
            w.writeheader()
            for inv in invs:
                cs = str(getattr(inv.overall_compliance_status, "value",
                                 inv.overall_compliance_status) or "unknown")
                w.writerow({
                    # ── ERP business columns ───────────────────────────────
                    "vendor_name":        inv.vendor_name or "",
                    "invoice_number":     inv.invoice_number or "",
                    "invoice_date":       inv.invoice_date or "",
                    "due_date":           inv.due_date or "",
                    "iban":               inv.iban or "",
                    "currency":           inv.currency or "CHF",
                    "total_amount":       f"{inv.total_amount:.2f}" if inv.total_amount else "",
                    "tax_amount":         f"{inv.tax_amount:.2f}" if inv.tax_amount else "",
                    "payment_reference":  inv.qr_reference or inv.invoice_number or "",
                    "compliance_status":  cs,
                    "anomaly_score":      inv.anomaly_score or 0,
                    # ── SAP FI/MM columns ──────────────────────────────────
                    "COMPANY_CODE":       "1000",
                    "DOC_TYPE":           "RE",
                    "PSTNG_DATE":         datetime.now().strftime("%Y%m%d"),
                    "PMNTTRMS":           inv.payment_terms or "",
                    "REF_DOC_NO":         inv.invoice_number or "",
                    "GROSS_AMOUNT":       f"{inv.total_amount:.2f}" if inv.total_amount else "",
                    "CALC_TAX_IND":       "X",
                    "TAX_AMOUNT":         f"{inv.tax_amount:.2f}" if inv.tax_amount else "",
                    "ITEM_TEXT":          f"Invoice {inv.invoice_number} from {inv.vendor_name}",
                    "VENDOR_NAME_SAP":    inv.vendor_name or "",
                    "IBAN_SAP":           inv.iban or "",
                    "VAT_REG_NO":         inv.vat_number or "",
                    "SWISS_UID":          inv.swiss_uid or "",
                    "QR_REFERENCE":       inv.qr_reference or "",
                })

        logger.info(f"SAP CSV saved: {path}")
        return path

    # ─────────────────────────────────────────────────────────────────────────
    # PUBLIC: Power BI Excel export (4 dashboard pages)
    # ─────────────────────────────────────────────────────────────────────────

    def export_powerbi_excel(self, invoice_ids=None, filename=None) -> Path:
        """
        Power BI-ready Excel workbook — 4 sheets matching dashboard pages:

        Page 1 — Executive Overview
          KPIs: total invoices, total value, by language, by currency,
          compliance pass/fail rate, anomaly count

        Page 2 — Compliance Monitoring
          Per-rule breakdown: fail/warning/pass counts, missing VAT/UID,
          missing IBAN, missing due date, invalid tax, high-risk invoices

        Page 3 — Vendor Risk
          Vendors ranked by anomaly score, IBAN risk flags,
          duplicate invoice numbers, high-value vendors

        Page 4 — Manual Review Queue
          invoice_id, vendor, issue, severity, recommended action
        """
        from openpyxl import Workbook
        from collections import Counter

        invs = self._invoices(invoice_ids)
        if not invs:
            raise ValueError("No invoices found.")
        ids = [i.id for i in invs]

        filename = filename or f"powerbi_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        path = settings.export_dir / filename
        wb = Workbook()

        # ── PAGE 1: Executive Overview ────────────────────────────────────
        ws1 = wb.active
        ws1.title = "Executive Overview"

        total     = len(invs)
        total_val = sum(i.total_amount or 0 for i in invs)
        pass_n    = sum(1 for i in invs if str(getattr(i.overall_compliance_status, "value",
                        i.overall_compliance_status)) == "pass")
        fail_n    = sum(1 for i in invs if str(getattr(i.overall_compliance_status, "value",
                        i.overall_compliance_status)) == "fail")
        warn_n    = sum(1 for i in invs if str(getattr(i.overall_compliance_status, "value",
                        i.overall_compliance_status)) == "warning")
        lang_cnt  = Counter(i.language.value if i.language else "unknown" for i in invs)
        curr_cnt  = Counter(i.currency or "—" for i in invs)
        anom_flags= self.db.query(AnomalyFlag).filter(AnomalyFlag.invoice_id.in_(ids)).count()

        exec_rows = [
            ("── KPIs ──", ""),
            ("Total Invoices Processed",     total),
            ("Total Invoice Value",          round(total_val, 2)),
            ("Compliance Pass",              pass_n),
            ("Compliance Warning",           warn_n),
            ("Compliance Fail",              fail_n),
            ("Pass Rate (%)",                round(pass_n / total * 100, 1) if total else 0),
            ("Total Anomaly Flags",          anom_flags),
            ("High-Risk Invoices (score≥60)", sum(1 for i in invs if (i.anomaly_score or 0) >= 60)),
            ("", ""),
            ("── By Language ──", ""),
        ] + [(f"  {lang}", cnt) for lang, cnt in lang_cnt.most_common()] + [
            ("", ""),
            ("── By Currency ──", ""),
        ] + [(f"  {curr}", cnt) for curr, cnt in curr_cnt.most_common()]

        self._kv_sheet(ws1, "Executive Overview", exec_rows)

        # ── PAGE 2: Compliance Monitoring ────────────────────────────────
        ws2 = wb.create_sheet("Compliance Monitoring")
        all_results = (
            self.db.query(ComplianceResult)
            .filter(ComplianceResult.invoice_id.in_(ids))
            .all()
        )
        rule_stats: dict[str, dict] = {}
        for r in all_results:
            rid = r.rule_id
            if rid not in rule_stats:
                rule_stats[rid] = {"Rule ID": rid, "Rule Name": r.rule_name,
                                   "Category": r.category,
                                   "Pass": 0, "Warning": 0, "Fail": 0}
            sv = r.status.value if r.status else "unknown"
            if sv == "pass":    rule_stats[rid]["Pass"]    += 1
            elif sv == "warning": rule_stats[rid]["Warning"] += 1
            elif sv == "fail":  rule_stats[rid]["Fail"]    += 1

        for v in rule_stats.values():
            total_checks = v["Pass"] + v["Warning"] + v["Fail"]
            v["Total Checks"] = total_checks
            v["Fail Rate (%)"] = round(v["Fail"] / total_checks * 100, 1) if total_checks else 0

        comp_df = pd.DataFrame(list(rule_stats.values()),
                               columns=["Rule ID","Rule Name","Category",
                                        "Pass","Warning","Fail","Total Checks","Fail Rate (%)"])
        comp_df.sort_values("Fail Rate (%)", ascending=False, inplace=True)

        if comp_df.empty:
            ws2["A1"] = "No compliance data."
        else:
            self._write_sheet(ws2, comp_df, header_color=_BLUE_DARK)

        # ── PAGE 3: Vendor Risk ───────────────────────────────────────────
        ws3 = wb.create_sheet("Vendor Risk")

        # Detect duplicate invoice numbers across vendors
        inv_num_map: dict[str, list[str]] = defaultdict(list)
        for inv in invs:
            if inv.invoice_number:
                inv_num_map[inv.invoice_number].append(inv.vendor_name or "?")

        vendor_rows = []
        for inv in invs:
            name = inv.vendor_name or "Unknown"
            dup_num = inv.invoice_number and len(inv_num_map.get(inv.invoice_number, [])) > 1
            vendor_rows.append({
                "Vendor Name":          name,
                "Invoice ID":           inv.id,
                "Invoice Number":       inv.invoice_number,
                "Total Amount":         inv.total_amount,
                "Currency":             inv.currency,
                "IBAN":                 inv.iban,
                "Swiss UID":            inv.swiss_uid,
                "Anomaly Score":        inv.anomaly_score or 0,
                "Compliance Status":    str(getattr(inv.overall_compliance_status, "value",
                                            inv.overall_compliance_status)),
                "Duplicate Invoice #":  "⚠️ YES" if dup_num else "No",
                "Risk Flag":            ("🔴 High" if (inv.anomaly_score or 0) >= 60
                                         else "🟡 Medium" if (inv.anomaly_score or 0) >= 30
                                         else "🟢 Low"),
            })

        vendor_df = pd.DataFrame(vendor_rows)
        vendor_df.sort_values("Anomaly Score", ascending=False, inplace=True)

        if vendor_df.empty:
            ws3["A1"] = "No vendor data."
        else:
            self._write_sheet(ws3, vendor_df, header_color=_BLUE_MID,
                              status_col="Compliance Status")

        # ── PAGE 4: Manual Review Queue ──────────────────────────────────
        ws4 = wb.create_sheet("Manual Review Queue")
        review_df = self._manual_review_df(invs)

        if review_df.empty:
            ws4["A1"] = "✅ No invoices require manual review."
        else:
            self._write_sheet(ws4, review_df, header_color=_BLUE_DARK,
                              status_col="Compliance Status")

        wb.save(path)
        logger.info(f"Power BI Excel saved: {path}")
        return path

    # ─────────────────────────────────────────────────────────────────────────
    # PUBLIC: Power BI JSON (legacy / kept for backward compatibility)
    # ─────────────────────────────────────────────────────────────────────────

    def export_powerbi_json(self, invoice_ids=None, filename=None) -> Path:
        """Denormalised JSON — all invoice data including compliance + anomaly flags."""
        invs = self._invoices(invoice_ids)
        if not invs:
            raise ValueError("No invoices found.")

        filename = filename or f"powerbi_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        path = settings.export_dir / filename

        records = []
        for inv in invs:
            crs = [{"rule_id": r.rule_id, "rule_name": r.rule_name,
                    "status": r.status.value if r.status else None,
                    "message": r.message, "category": r.category} for r in inv.compliance_results]
            flags = [{"anomaly_type": f.anomaly_type, "severity": f.severity,
                      "score_contribution": f.score_contribution,
                      "description": f.description,
                      "recommended_action": f.recommended_action} for f in inv.anomaly_flags]
            lis = [{"pos": l.position, "desc": l.description,
                    "qty": l.quantity, "total": l.total_price} for l in inv.line_items]
            statuses = [c["status"] for c in crs]
            overall  = ("fail" if "fail" in statuses else
                        "warning" if "warning" in statuses else
                        "pass" if statuses else "unknown")
            records.append({
                "invoice_id": inv.id, "filename": inv.original_filename,
                "invoice_number": inv.invoice_number, "vendor_name": inv.vendor_name,
                "currency": inv.currency, "total_amount": inv.total_amount,
                "tax_amount": inv.tax_amount, "invoice_date": inv.invoice_date,
                "due_date": inv.due_date, "iban": inv.iban,
                "swiss_uid": inv.swiss_uid, "qr_reference": inv.qr_reference,
                "language": inv.language.value if inv.language else None,
                "anomaly_score": inv.anomaly_score or 0,
                "overall_compliance": overall,
                "fail_count": statuses.count("fail"),
                "warning_count": statuses.count("warning"),
                "compliance_results": crs,
                "anomaly_flags": flags,
                "line_items": lis,
            })

        payload = {
            "metadata": {
                "generated_at": datetime.utcnow().isoformat() + "Z",
                "total_records": len(records),
                "schema_version": "1.2",
            },
            "invoices": records,
        }
        with open(path, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)

        logger.info(f"Power BI JSON saved: {path}")
        return path
