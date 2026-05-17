"""
Tests for ExportService — v1.2
Covers: export_excel, export_sap_csv, export_powerbi_excel, export_powerbi_json
"""
import csv
import json
import os
from datetime import datetime
from pathlib import Path

import pytest

from app.models.invoice import (
    Invoice, ComplianceResult, AnomalyFlag,
    ProcessingStatus, ComplianceStatus, InvoiceLanguage,
)
from app.services.export_service import ExportService


# ── Helpers ───────────────────────────────────────────────────────────────────

def _make_invoice(db, **kwargs) -> Invoice:
    defaults = dict(
        original_filename="test.pdf",
        file_path="/tmp/test.pdf",
        file_type="pdf",
        status=ProcessingStatus.COMPLETED,
        invoice_number="INV-001",
        vendor_name="Mustermann AG",
        currency="CHF",
        total_amount=3729.45,
        tax_amount=279.45,
        tax_rate_percent=8.1,
        invoice_date="2024-03-15",
        due_date="2024-04-15",
        swiss_uid="CHE-123.456.789",
        iban="CH56048350123456780090",
        language=InvoiceLanguage.DE,
        extraction_confidence=0.95,
        anomaly_score=0,
    )
    defaults.update(kwargs)
    inv = Invoice(**defaults)
    db.add(inv)
    db.flush()
    return inv


def _add_compliance(db, invoice_id: int, status: str = "pass",
                    rule_id: str = "CH_UID_FORMAT"):
    cr = ComplianceResult(
        invoice_id=invoice_id,
        rule_id=rule_id,
        rule_name="Test Rule",
        category="identity",
        status=ComplianceStatus[status.upper()],
        message=f"Rule {status}",
        field_checked="swiss_uid",
        actual_value="CHE-123.456.789",
    )
    db.add(cr)
    db.flush()
    return cr


def _add_anomaly(db, invoice_id: int) -> AnomalyFlag:
    flag = AnomalyFlag(
        invoice_id=invoice_id,
        anomaly_type="duplicate_invoice_number",
        severity="high",
        score_contribution=40,
        description="Duplicate invoice number detected.",
        recommended_action="Verify with vendor",
    )
    db.add(flag)
    db.flush()
    return flag


def _svc(db) -> ExportService:
    os.makedirs("/tmp/test_exports", exist_ok=True)
    return ExportService(db)


# ── export_excel ──────────────────────────────────────────────────────────────

class TestExportExcel:
    def test_creates_xlsx_file(self, db):
        _make_invoice(db)
        db.commit()
        path = _svc(db).export_excel()
        assert path.exists()
        assert path.suffix == ".xlsx"

    def test_excel_has_required_sheets(self, db):
        import openpyxl
        inv = _make_invoice(db)
        _add_compliance(db, inv.id, "fail")
        _add_anomaly(db, inv.id)
        db.commit()
        path = _svc(db).export_excel()
        wb = openpyxl.load_workbook(path)
        for sheet in ("Summary", "Invoices", "Vendors", "Compliance Issues",
                      "Anomalies", "Manual Review", "Line Items"):
            assert sheet in wb.sheetnames, f"Missing sheet: {sheet}"

    def test_invoices_sheet_has_data(self, db):
        import openpyxl
        _make_invoice(db, vendor_name="Acme GmbH")
        db.commit()
        path = _svc(db).export_excel()
        wb   = openpyxl.load_workbook(path)
        ws   = wb["Invoices"]
        # row 1 = header, row 2 = first data row
        values = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
        assert any("Acme GmbH" in str(v) for v in values)

    def test_no_invoices_raises(self, db):
        with pytest.raises(ValueError, match="No invoices found"):
            _svc(db).export_excel()

    def test_manual_review_sheet_populated_for_high_score(self, db):
        import openpyxl
        inv = _make_invoice(db, anomaly_score=75)
        _add_compliance(db, inv.id, "fail")
        db.commit()
        path = _svc(db).export_excel()
        wb   = openpyxl.load_workbook(path)
        ws   = wb["Manual Review"]
        # Should have at least 1 data row
        assert ws.max_row >= 2

    def test_vendor_sheet_aggregates_correctly(self, db):
        import openpyxl
        _make_invoice(db, vendor_name="BigCorp", total_amount=1000.0)
        _make_invoice(db, vendor_name="BigCorp", total_amount=2000.0,
                      invoice_number="INV-002")
        db.commit()
        path = _svc(db).export_excel()
        wb   = openpyxl.load_workbook(path)
        ws   = wb["Vendors"]
        # Find BigCorp row
        names = [ws.cell(r, 1).value for r in range(2, ws.max_row + 1)]
        assert "BigCorp" in names


# ── export_sap_csv ────────────────────────────────────────────────────────────

class TestExportSapCsv:
    def test_creates_csv_file(self, db):
        _make_invoice(db)
        db.commit()
        path = _svc(db).export_sap_csv()
        assert path.exists()
        assert path.suffix == ".csv"

    def test_csv_has_required_columns(self, db):
        _make_invoice(db)
        db.commit()
        path = _svc(db).export_sap_csv()
        with open(path, encoding="utf-8-sig") as f:
            reader = csv.DictReader(f, delimiter=";")
            headers = reader.fieldnames or []
        required = [
            "vendor_name", "invoice_number", "invoice_date", "due_date",
            "iban", "currency", "total_amount", "tax_amount",
            "payment_reference", "compliance_status", "anomaly_score",
        ]
        for col in required:
            assert col in headers, f"Missing column: {col}"

    def test_csv_values_correct(self, db):
        _make_invoice(db, vendor_name="Testfirma AG", invoice_number="SAP-001",
                      currency="EUR", total_amount=500.0, anomaly_score=20)
        db.commit()
        path = _svc(db).export_sap_csv()
        with open(path, encoding="utf-8-sig") as f:
            reader = csv.DictReader(f, delimiter=";")
            row = next(reader)
        assert row["vendor_name"] == "Testfirma AG"
        assert row["invoice_number"] == "SAP-001"
        assert row["currency"] == "EUR"
        assert row["total_amount"] == "500.00"
        assert row["anomaly_score"] == "20"

    def test_no_invoices_raises(self, db):
        with pytest.raises(ValueError):
            _svc(db).export_sap_csv()


# ── export_powerbi_excel ──────────────────────────────────────────────────────

class TestExportPowerBiExcel:
    def test_creates_xlsx_file(self, db):
        _make_invoice(db)
        db.commit()
        path = _svc(db).export_powerbi_excel()
        assert path.exists()
        assert path.suffix == ".xlsx"

    def test_has_four_sheets(self, db):
        import openpyxl
        _make_invoice(db)
        db.commit()
        path = _svc(db).export_powerbi_excel()
        wb   = openpyxl.load_workbook(path)
        for sheet in ("Executive Overview", "Compliance Monitoring",
                      "Vendor Risk", "Manual Review Queue"):
            assert sheet in wb.sheetnames, f"Missing sheet: {sheet}"

    def test_executive_overview_has_kpis(self, db):
        import openpyxl
        _make_invoice(db)
        db.commit()
        path = _svc(db).export_powerbi_excel()
        wb   = openpyxl.load_workbook(path)
        ws   = wb["Executive Overview"]
        labels = [ws.cell(r, 1).value for r in range(2, ws.max_row + 1)]
        assert "Total Invoices Processed" in labels
        assert "Total Invoice Value" in labels
        assert "Pass Rate (%)" in labels

    def test_vendor_risk_sorted_by_anomaly_score(self, db):
        import openpyxl
        _make_invoice(db, vendor_name="LowRisk",  anomaly_score=5)
        _make_invoice(db, vendor_name="HighRisk", anomaly_score=80,
                      invoice_number="INV-HR-001")
        db.commit()
        path = _svc(db).export_powerbi_excel()
        wb   = openpyxl.load_workbook(path)
        ws   = wb["Vendor Risk"]
        # First data row should be the highest-risk vendor
        first_vendor = ws.cell(2, 1).value
        assert first_vendor == "HighRisk"

    def test_no_invoices_raises(self, db):
        with pytest.raises(ValueError):
            _svc(db).export_powerbi_excel()


# ── export_powerbi_json ───────────────────────────────────────────────────────

class TestExportPowerBiJson:
    def test_creates_json_file(self, db):
        _make_invoice(db)
        db.commit()
        path = _svc(db).export_powerbi_json()
        assert path.exists()
        assert path.suffix == ".json"

    def test_json_structure(self, db):
        inv = _make_invoice(db)
        _add_compliance(db, inv.id, "pass")
        db.commit()
        path = _svc(db).export_powerbi_json()
        with open(path) as f:
            data = json.load(f)
        assert "metadata" in data
        assert "invoices" in data
        assert data["metadata"]["schema_version"] == "1.2"
        assert len(data["invoices"]) == 1

    def test_json_includes_compliance_and_anomaly(self, db):
        inv = _make_invoice(db)
        _add_compliance(db, inv.id, "fail", "CH_MISSING_UID")
        _add_anomaly(db, inv.id)
        db.commit()
        path = _svc(db).export_powerbi_json()
        with open(path) as f:
            data = json.load(f)
        rec = data["invoices"][0]
        assert len(rec["compliance_results"]) == 1
        assert len(rec["anomaly_flags"]) == 1
        assert rec["anomaly_flags"][0]["anomaly_type"] == "duplicate_invoice_number"

    def test_no_invoices_raises(self, db):
        with pytest.raises(ValueError):
            _svc(db).export_powerbi_json()


# ── Currency acceptance ───────────────────────────────────────────────────────

class TestCurrencyConfig:
    def test_accepts_major_fiat_currencies(self):
        from app.config import settings
        currencies = settings.accepted_currencies_list
        for c in ["CHF", "EUR", "USD", "GBP", "CAD", "JPY", "SGD", "AUD",
                   "CNY", "KRW", "INR", "BRL", "NOK", "SEK", "DKK", "NZD"]:
            assert c in currencies, f"Missing currency: {c}"

    def test_accepts_crypto_currencies(self):
        from app.config import settings
        currencies = settings.accepted_currencies_list
        for c in ["BTC", "ETH", "USDT", "BNB", "XRP"]:
            assert c in currencies, f"Missing crypto: {c}"

    def test_accepts_ruble_and_others(self):
        from app.config import settings
        currencies = settings.accepted_currencies_list
        for c in ["RUB", "ZAR", "TRY", "THB", "PLN", "SAR", "MYR", "MXN"]:
            assert c in currencies, f"Missing currency: {c}"
